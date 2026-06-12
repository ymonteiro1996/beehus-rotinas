from flask import Blueprint, render_template, request, send_file, jsonify
from db import db, get_company_filter
import io, os, json, uuid, openpyxl, smtplib
from bisect import bisect_left, bisect_right
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bson import ObjectId

SAVED_REPORTS_FILE  = os.path.join(os.path.dirname(__file__), "..", "data", "saved_reports.json")
WALLET_TMPL_FILE   = os.path.join(os.path.dirname(__file__), "..", "data", "report_wallet_templates.json")
DERIV_MAPPINGS_FILE = os.path.join(os.path.dirname(__file__), "..", "data", "deriv_mappings.json")


def _load_saved():
    if not os.path.exists(SAVED_REPORTS_FILE):
        return []
    with open(SAVED_REPORTS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _write_saved(reports):
    with open(SAVED_REPORTS_FILE, "w", encoding="utf-8") as f:
        json.dump(reports, f, indent=2, ensure_ascii=False)


def _load_wallet_templates():
    if not os.path.exists(WALLET_TMPL_FILE):
        return []
    with open(WALLET_TMPL_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _write_wallet_templates(templates):
    with open(WALLET_TMPL_FILE, "w", encoding="utf-8") as f:
        json.dump(templates, f, indent=2, ensure_ascii=False)

bp = Blueprint("report_builder", __name__)

# ══════════════════════════════════════════════════════════════════════════════
# COLLECTION & FIELD DEFINITIONS
# type: date | str | num | ref
# nested=True  → field lives inside securities[] (processedPosition only)
# hv           → key inside hierarchicalVariable sub-object
# ref + name_field → resolved from lookup map when resolve_refs=True
# ══════════════════════════════════════════════════════════════════════════════

COLLECTIONS = {
    "processedPosition": {
        "label": "Posições Processadas",
        "date_field": "positionDate",
        "company_field": "companyId",
        "entity_field": "entityId",
        "wallet_field": "walletId",
        "fields": [
            {"key": "positionDate",      "label": "Data",                "type": "date"},
            {"key": "companyId",         "label": "Empresa",             "type": "ref",  "ref": "companies",  "name_field": "name"},
            {"key": "entityId",          "label": "Entidade",            "type": "ref",  "ref": "entities",   "name_field": "name"},
            {"key": "walletId",          "label": "Carteira",            "type": "ref",  "ref": "wallets",    "name_field": "name"},
            {"key": "beehusName",        "label": "Ativo",               "type": "str",  "nested": True},
            {"key": "quantity",          "label": "Quantidade",          "type": "num",  "nested": True},
            {"key": "pu",                "label": "PU",                  "type": "num",  "nested": True},
            {"key": "amount",            "label": "Saldo",               "type": "num",  "nested": True},
            {"key": "pricingType",       "label": "Precificação",        "type": "str",  "nested": True},
            {"key": "variable1",         "label": "Classe",              "type": "str",  "nested": True, "hv": "variable1"},
            {"key": "variable2",         "label": "Subclasse",           "type": "str",  "nested": True, "hv": "variable2"},
            {"key": "dailyContribution", "label": "Contribuição diária", "type": "num",  "nested": True},
        ],
    },
    "transactions": {
        "label": "Transações",
        "date_field": "operationDate",
        "company_field": "companyId",
        "wallet_field": "walletId",
        "security_id_field": "securityId",
        "fields": [
            {"key": "operationDate",         "label": "Data Operação",   "type": "date"},
            {"key": "liquidationDate",       "label": "Data Liquidação", "type": "date"},
            {"key": "companyId",             "label": "Empresa",         "type": "ref",  "ref": "companies",  "name_field": "name"},
            {"key": "walletId",              "label": "Carteira",        "type": "ref",  "ref": "wallets",    "name_field": "name"},
            {"key": "securityId",            "label": "Ativo",           "type": "ref",  "ref": "securities", "name_field": "beehusName"},
            {"key": "beehusTransactionType", "label": "Tipo",            "type": "str"},
            {"key": "quantity",              "label": "Quantidade",      "type": "num"},
            {"key": "price",                 "label": "Preço",           "type": "num"},
            {"key": "balance",               "label": "Valor",           "type": "num"},
            {"key": "description",           "label": "Descrição",       "type": "str"},
        ],
    },
    "provisions": {
        "label": "Provisões",
        "date_field": "initialDate",
        "company_field": "companyId",
        "wallet_field": "walletId",
        "security_id_field": "securityId",
        "fields": [
            {"key": "initialDate",     "label": "Data Inicial",    "type": "date"},
            {"key": "liquidationDate", "label": "Data Liquidação", "type": "date"},
            {"key": "companyId",       "label": "Empresa",         "type": "ref",  "ref": "companies",  "name_field": "name"},
            {"key": "walletId",        "label": "Carteira",        "type": "ref",  "ref": "wallets",    "name_field": "name"},
            {"key": "securityId",      "label": "Ativo",           "type": "ref",  "ref": "securities", "name_field": "beehusName"},
            {"key": "provisionType",   "label": "Tipo",            "type": "str"},
            {"key": "balance",         "label": "Valor",           "type": "num"},
            {"key": "description",     "label": "Descrição",       "type": "str"},
        ],
    },
    "publishedPositionSecurities": {
        "label": "Posições Publicadas",
        "date_field": "positionDate",
        "company_field": "companyId",
        "entity_field": "entityId",
        "wallet_field": "walletId",
        "security_id_field": "securityId",
        "fields": [
            {"key": "positionDate",      "label": "Data",                "type": "date"},
            {"key": "companyId",         "label": "Empresa",             "type": "ref",  "ref": "companies",  "name_field": "name"},
            {"key": "entityId",          "label": "Entidade",            "type": "ref",  "ref": "entities",   "name_field": "name"},
            {"key": "walletId",          "label": "Carteira",            "type": "ref",  "ref": "wallets",    "name_field": "name"},
            {"key": "securityId",        "label": "Ativo",               "type": "ref",  "ref": "securities", "name_field": "beehusName"},
            {"key": "quantity",          "label": "Quantidade",          "type": "num"},
            {"key": "pu",                "label": "PU",                  "type": "num"},
            {"key": "amount",            "label": "Saldo",               "type": "num"},
            {"key": "pricingType",       "label": "Precificação",        "type": "str"},
            {"key": "variable1",         "label": "Classe",              "type": "str"},
            {"key": "variable2",         "label": "Subclasse",           "type": "str"},
            {"key": "dailyContribution", "label": "Contribuição diária", "type": "num"},
        ],
    },
    "cashAccounts": {
        "label": "Contas Correntes",
        "date_field": "positionDate",
        "company_field": "companyId",
        "wallet_field": "walletId",
        "fields": [
            {"key": "positionDate", "label": "Data",        "type": "date"},
            {"key": "companyId",    "label": "Empresa",     "type": "ref", "ref": "companies", "name_field": "name"},
            {"key": "walletId",     "label": "Carteira",    "type": "ref", "ref": "wallets",   "name_field": "name"},
            {"key": "balance",      "label": "Saldo",       "type": "num"},
            {"key": "currency",     "label": "Moeda",       "type": "str"},
            {"key": "accountCode",  "label": "Conta",       "type": "str"},
            {"key": "description",  "label": "Descrição",   "type": "str"},
        ],
    },
    "navPackages": {
        "label": "NAV Packages",
        "date_field": "positionDate",
        "company_field": "companyId",
        "wallet_field": "walletId",
        "no_securities": True,
        "fields": [
            {"key": "positionDate",       "label": "Data",                   "type": "date"},
            {"key": "companyId",          "label": "Empresa",                "type": "ref",  "ref": "companies", "name_field": "name"},
            {"key": "walletId",           "label": "Carteira",               "type": "ref",  "ref": "wallets",   "name_field": "name"},
            {"key": "nav",                "label": "NAV",                    "type": "num"},
            {"key": "navPerShare",        "label": "NAV por Cota",           "type": "num"},
            {"key": "amount",             "label": "Patrimônio",             "type": "num"},
            {"key": "formerAmount",       "label": "Patrimônio Anterior",    "type": "num"},
            {"key": "inAndOutFlows",      "label": "inAndOutFlows",          "type": "num"},
            {"key": "returnNavPerShare",  "label": "Retorno NAV/Cota",       "type": "num"},
            {"key": "returnContribution", "label": "Retorno Contribuição",   "type": "num"},
            {"key": "currency",           "label": "Moeda",                  "type": "str"},
            {"key": "published",          "label": "Publicado",              "type": "str"},
        ],
    },
    "unprocessedSecurityPositions": {
        "label": "Posições Não Processadas",
        "date_field": "positionDate",
        "company_field": "companyId",
        "wallet_field": "walletId",
        "mapping_via": "unprocessedId",   # field in securities[] matched via securityMappings
        "fields": [
            {"key": "positionDate",  "label": "Data",        "type": "date"},
            {"key": "companyId",     "label": "Empresa",     "type": "ref", "ref": "companies", "name_field": "name"},
            {"key": "walletId",      "label": "Carteira",    "type": "ref", "ref": "wallets",   "name_field": "name"},
            {"key": "unprocessedId", "label": "ID Externo",  "type": "str", "nested": True},
            {"key": "securityId",    "label": "Ativo",       "type": "ref", "ref": "securities", "name_field": "beehusName", "nested": True, "mapped": True},
            {"key": "quantity",      "label": "Quantidade",  "type": "num", "nested": True},
            {"key": "pu",            "label": "PU",          "type": "num", "nested": True},
            {"key": "amount",        "label": "Saldo",       "type": "num", "nested": True},
        ],
    },
}

# Extra fields sourced from the securities collection (joined via securityId)
_SEC_FIELDS = [
    {"key": "sec_mainId",            "label": "Código Principal",  "type": "str",  "source": "securities", "sec_field": "mainId"},
    {"key": "sec_maturityDate",      "label": "Vencimento",        "type": "date", "source": "securities", "sec_field": "maturityDate"},
    {"key": "sec_yield",             "label": "Taxa",              "type": "num",  "source": "securities", "sec_field": "yield"},
    {"key": "sec_indexer",           "label": "Indexador",         "type": "str",  "source": "securities", "sec_field": "indexer"},
    {"key": "sec_indexerPercentual", "label": "% do Indexador",    "type": "num",  "source": "securities", "sec_field": "indexerPercentual"},
    {"key": "sec_securityType",      "label": "Tipo de Ativo",     "type": "str",  "source": "securities", "sec_field": "securityType"},
    {"key": "sec_type",              "label": "Sub-tipo",          "type": "str",  "source": "securities", "sec_field": "type"},
]

# Inject into every collection that can link to a security document
_NESTED_COLLS = {"processedPosition", "unprocessedSecurityPositions"}
for _k, _c in COLLECTIONS.items():
    if _c.get("security_id_field") or _k in _NESTED_COLLS:
        _nested = _k in _NESTED_COLLS
        _c["fields"] += [{**f, "nested": True} if _nested else f for f in _SEC_FIELDS]

# Calculated columns available per collection
_NAV_RENT_CALC = {"key": "calc_nav_rent", "label": "Rentabilidade",
                  "formula": "(navPerShare final ÷ navPerShare inicial) − 1",
                  "period_based": True, "source": "navPerShare"}

_PU_RENT_CALC = {"key": "calc_pu_rent", "label": "Rentabilidade PU",
                 "formula": "(PU final ÷ PU inicial) − 1",
                 "period_based": True, "source": "pu"}

_CONTRIB_RENT_CALC = {"key": "calc_contrib_rent", "label": "Rentabilidade Contribuição",
                      "formula": "Π(1 + totalContribution / saldoAnterior) − 1",
                      "period_based": True, "source": "contribution"}

_CALC_DEFS = {
    "navPackages": [_NAV_RENT_CALC],
    "processedPosition": [
        {"key": "calc_rent_pu",     "label": "Rentabilidade Diária (PU)",
         "formula": "(PU ÷ PU anterior) − 1"},
        {"key": "calc_rent_contrib","label": "Rentabilidade Diária (Contribuição)",
         "formula": "Contribuição Total ÷ (PU anterior × Qtd anterior)"},
        _PU_RENT_CALC,
        _CONTRIB_RENT_CALC,
    ],
    "publishedPositionSecurities": [
        {"key": "calc_rent_pu",     "label": "Rentabilidade Diária (PU)",
         "formula": "(PU ÷ PU anterior) − 1"},
        {"key": "calc_rent_contrib","label": "Rentabilidade Diária (Contribuição)",
         "formula": "Contribuição Total ÷ (PU anterior × Qtd anterior)"},
        _PU_RENT_CALC,
        _CONTRIB_RENT_CALC,
    ],
    "unprocessedSecurityPositions": [_PU_RENT_CALC],
}


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _to_oid(v):
    try:
        return ObjectId(v)
    except Exception:
        return v


def _get_lookup_maps():
    _sec_proj = {"mainId": 1, "maturityDate": 1, "yield": 1,
                 "indexer": 1, "indexerPercentual": 1, "securityType": 1, "beehusName": 1,
                 "type": 1}
    return {
        "companies":   {str(c["_id"]): c.get("name", "")       for c in db.companies.find({}, {"name": 1})},
        "entities":    {str(e["_id"]): e.get("name", "")       for e in db.entities.find({}, {"name": 1})},
        "wallets":     {str(w["_id"]): w.get("name", "")       for w in db.wallets.find({}, {"name": 1})},
        "securities":  {str(s["_id"]): s.get("beehusName", "") for s in db.securities.find({}, {"beehusName": 1})},
        "sec_details": {str(s["_id"]): s                       for s in db.securities.find({}, _sec_proj)},
    }


def _build_security_mapping(company_id):
    """Return {unprocessedId: securityId_str} from securityMappings for a given company."""
    result = {}
    doc = db.securityMappings.find_one({"companyId": company_id})
    if doc:
        for m in (doc.get("mappings") or []):
            from_id = m.get("from")
            sec_id  = m.get("securityId") or m.get("to")
            if from_id and sec_id:
                result[str(from_id)] = str(sec_id)
    return result


def _build_query(coll_config, company_id, entity_id, wallet_id, date_from, date_to, security_ids=None):
    q = {"trashed": {"$ne": True}}

    # All ID fields are stored as plain strings in all collections
    if company_id:
        q[coll_config["company_field"]] = company_id

    ef = coll_config.get("entity_field")
    if entity_id and ef:
        q[ef] = entity_id

    wf = coll_config.get("wallet_field")
    if wallet_id and wf:
        if isinstance(wallet_id, list):
            q[wf] = {"$in": wallet_id}
        else:
            q[wf] = wallet_id

    df = coll_config.get("date_field")
    if df:
        dq = {}
        if date_from: dq["$gte"] = date_from
        if date_to:   dq["$lte"] = date_to
        if dq:        q[df] = dq

    sf = coll_config.get("security_id_field")
    if security_ids and sf:
        q[sf] = {"$in": security_ids}

    return q


def _build_rows(coll_key, coll_config, selected_keys, query, lookup_maps, security_ids=None, need_calc_fields=False):
    """selected_keys may include '<key>__name' variants for ref fields.
    need_calc_fields: if True, also populate _calc_* hidden keys used by _apply_calcs."""
    field_map   = {f["key"]: f for f in coll_config["fields"]}
    sec_filter  = set(security_ids) if security_ids is not None else None
    sec_details = lookup_maps.get("sec_details", {})

    # Unique base keys needed (strip __name suffix)
    base_keys = list(dict.fromkeys(
        k[:-6] if k.endswith("__name") else k for k in selected_keys
    ))
    sel = [field_map[k] for k in base_keys if k in field_map]

    def _resolve_name(fd, val):
        return lookup_maps.get(fd["ref"], {}).get(str(val), str(val)) if val else ""

    def _get_sec(sec_id):
        return sec_details.get(str(sec_id), {}) if sec_id else {}

    def _fill(row, fd, raw_val, sec_doc=None):
        if fd.get("source") == "securities":
            row[fd["key"]] = (sec_doc or {}).get(fd.get("sec_field", ""))
            return
        for out_key in selected_keys:
            is_name = out_key.endswith("__name")
            base    = out_key[:-6] if is_name else out_key
            if base != fd["key"]:
                continue
            row[out_key] = _resolve_name(fd, raw_val) if is_name else raw_val

    rows = []

    if coll_key == "processedPosition":
        for doc in db[coll_key].find(query):
            base_row = {"_raw_walletId": str(doc.get("walletId", "") or "")}
            for fd in sel:
                if not fd.get("nested"):
                    _fill(base_row, fd, doc.get(fd["key"]))

            for sec in (doc.get("securities") or [{}]):
                if sec_filter is not None and str(sec.get("securityId", "")) not in sec_filter:
                    continue
                sec_doc = _get_sec(sec.get("securityId"))
                row = dict(base_row)
                row["_raw_securityId"] = str(sec.get("securityId", "") or "")
                for fd in sel:
                    if not fd.get("nested"):
                        continue
                    if fd.get("source") == "securities":
                        _fill(row, fd, None, sec_doc)
                    elif "hv" in fd:
                        hv = sec.get("hierarchicalVariable") or {}
                        row[fd["key"]] = hv.get(fd["hv"], "")
                    else:
                        row[fd["key"]] = sec.get(fd["key"])
                if need_calc_fields:
                    row["_calc_sec_id"]        = sec.get("securityId")
                    row["_calc_date"]          = doc.get("positionDate")
                    row["_calc_pu"]            = sec.get("pu")
                    row["_calc_qty"]           = sec.get("quantity")
                    row["_calc_total_contrib"] = sec.get("totalContribution")
                rows.append(row)

    elif coll_key == "unprocessedSecurityPositions":
        company_id  = query.get(coll_config.get("company_field", "companyId"), "")
        sec_mapping = _build_security_mapping(company_id)
        for doc in db[coll_key].find(query):
            base_row = {"_raw_walletId": str(doc.get("walletId", "") or "")}
            for fd in sel:
                if not fd.get("nested"):
                    _fill(base_row, fd, doc.get(fd["key"]))

            for sec in (doc.get("securities") or [{}]):
                unproc_id   = str(sec.get("unprocessedId", "") or "")
                sec_id      = sec_mapping.get(unproc_id)
                unmapped    = sec_id is None and bool(unproc_id)
                sec_doc     = _get_sec(sec_id) if sec_id else {}
                if sec_filter is not None and sec_id not in sec_filter:
                    continue
                row = dict(base_row)
                row["_raw_securityId"] = unproc_id
                for fd in sel:
                    if not fd.get("nested"):
                        continue
                    if fd.get("source") == "securities":
                        _fill(row, fd, None, sec_doc)
                    elif fd.get("mapped"):
                        # securityId resolved via securityMappings; flag if missing
                        for out_key in selected_keys:
                            is_name = out_key.endswith("__name")
                            base    = out_key[:-6] if is_name else out_key
                            if base != fd["key"]:
                                continue
                            if unmapped:
                                row[out_key] = "[sem mapeamento]"
                            else:
                                row[out_key] = _resolve_name(fd, sec_id) if is_name else sec_id
                    else:
                        _fill(row, fd, sec.get(fd["key"]))
                if need_calc_fields:
                    row["_calc_sec_id"]        = sec_id
                    row["_calc_date"]          = doc.get("positionDate")
                    row["_calc_pu"]            = sec.get("pu")
                    row["_calc_qty"]           = sec.get("quantity")
                    row["_calc_total_contrib"] = sec.get("totalContribution")
                rows.append(row)

    else:
        sid_field = coll_config.get("security_id_field")
        wf        = coll_config.get("wallet_field")
        for doc in db[coll_key].find(query):
            sec_doc = _get_sec(doc.get(sid_field)) if sid_field else {}
            row = {}
            for fd in sel:
                _fill(row, fd, doc.get(fd["key"]), sec_doc)
            if wf:
                row["_raw_walletId"] = str(doc.get(wf, "") or "")
            if sid_field:
                row["_raw_securityId"] = str(doc.get(sid_field, "") or "")
            if need_calc_fields:
                row["_calc_sec_id"]        = doc.get(sid_field or "securityId")
                row["_calc_date"]          = doc.get(coll_config["date_field"])
                row["_calc_pu"]            = doc.get("pu")
                row["_calc_qty"]           = doc.get("quantity")
                row["_calc_total_contrib"] = doc.get("totalContribution")
            rows.append(row)

    return rows


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════

_HDR_FILL = PatternFill("solid", fgColor="1A1A1A")
_ALT_FILL = PatternFill("solid", fgColor="F7F7F7")
_HDR_FONT = Font(color="FFFFFF", bold=True, size=9)
_BODY_FONT = Font(size=9)
_THIN = Side(style="thin", color="E0E0E0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)


def _write_sheet(ws, headers, row_keys, rows, num_keys, date_keys=None, pct_keys=None):
    date_keys = date_keys or set()
    pct_keys  = pct_keys  or set()
    ws.freeze_panes = "A2"

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = _HDR_FILL, _HDR_FONT, _CENTER, _BORDER

    for r_idx, row in enumerate(rows, 2):
        fill = _ALT_FILL if r_idx % 2 == 0 else None
        for col, key in enumerate(row_keys, 1):
            val = row.get(key)
            if key in date_keys and isinstance(val, str) and len(val) >= 10:
                val = val[8:10] + "/" + val[5:7] + "/" + val[:4]
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.font      = _BODY_FONT
            cell.alignment = _CENTER if key in num_keys else _LEFT
            cell.border    = _BORDER
            if fill:
                cell.fill = fill
            if key in pct_keys and isinstance(val, (int, float)):
                cell.number_format = "0.0000%"
            elif key in num_keys and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"

    for col in ws.columns:
        length = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, 10), 45)


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@bp.route("/report-builder")
def index():
    companies = sorted(
        [{"id": str(c["_id"]), "name": c.get("name", "")} for c in db.companies.find({}, {"name": 1})],
        key=lambda c: c["name"],
    )
    cf = get_company_filter()
    if cf:
        companies = [c for c in companies if c["id"] in cf]
    sorted_collections = dict(sorted(COLLECTIONS.items(), key=lambda x: x[1]["label"]))
    return render_template("report_builder.html", companies=companies, collections=sorted_collections)


@bp.route("/api/report-builder/wallets")
def get_wallets():
    company_id = request.args.get("companyId", "").strip()
    if not company_id:
        return jsonify([])
    wallets = list(db.wallets.find(
        {"companyId": company_id},
        {"name": 1, "accountCode": 1}
    ).sort("name", 1))
    return jsonify([
        {"id": str(w["_id"]), "name": w.get("name", ""), "accountCode": w.get("accountCode", "")}
        for w in wallets
    ])


@bp.route("/api/report-builder/wallet-templates")
def get_wallet_tmpl():
    return jsonify(_load_wallet_templates())


@bp.route("/api/report-builder/wallet-templates", methods=["POST"])
def save_wallet_tmpl():
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Nome obrigatório"}), 400
    template = {"name": name, "wallets": data.get("wallets", [])}
    templates = _load_wallet_templates()
    for t in templates:
        if t.get("name") == name:
            t["wallets"] = template["wallets"]
            _write_wallet_templates(templates)
            return jsonify({"ok": True})
    templates.append(template)
    _write_wallet_templates(templates)
    return jsonify({"ok": True})


@bp.route("/api/report-builder/wallet-templates/<name>", methods=["DELETE"])
def delete_wallet_tmpl(name):
    templates = _load_wallet_templates()
    _write_wallet_templates([t for t in templates if t.get("name") != name])
    return jsonify({"ok": True})


@bp.route("/api/report-builder/filters")
def get_filters():
    company_id = request.args.get("company_id", "")

    # wallets store companyId/entityId as plain strings
    entity_id = request.args.get("entity_id", "")
    wallet_q = {}
    if company_id: wallet_q["companyId"] = company_id
    if entity_id:  wallet_q["entityId"]  = entity_id
    company_wallets = list(db.wallets.find(wallet_q, {"name": 1, "entityId": 1}))

    # derive entity IDs from wallets of that company
    entity_ids = {w.get("entityId") for w in company_wallets if w.get("entityId")}
    entity_q   = {"_id": {"$in": [_to_oid(eid) for eid in entity_ids]}} if entity_ids else {"_id": {"$in": []}}

    entities = sorted(
        [{"id": str(e["_id"]), "name": e.get("name", "")} for e in db.entities.find(entity_q, {"name": 1})],
        key=lambda x: x["name"],
    )
    wallets = sorted(
        [{"id": str(w["_id"]), "name": w.get("name", "")} for w in company_wallets],
        key=lambda x: x["name"],
    )
    return jsonify({"entities": entities, "wallets": wallets})


@bp.route("/api/report-builder/security-types")
def get_security_types():
    types = sorted(t for t in db.securities.distinct("securityType") if t)
    return jsonify(types)


@bp.route("/api/report-builder/date-range")
def get_date_range():
    company_id = request.args.get("company_id", "")
    entity_id  = request.args.get("entity_id",  "")
    wallet_ids = request.args.getlist("wallet_ids")
    wallet_id  = wallet_ids if wallet_ids else request.args.get("wallet_id", "")
    coll_keys  = request.args.getlist("collections")

    result = {}
    for coll_key in coll_keys:
        if coll_key not in COLLECTIONS:
            continue
        cfg = COLLECTIONS[coll_key]
        df  = cfg.get("date_field")
        if not df:
            continue
        q = _build_query(cfg, company_id, entity_id, wallet_id, "", "")
        proj = {df: 1}
        first = db[coll_key].find_one(q, proj, sort=[(df, 1)])
        last  = db[coll_key].find_one(q, proj, sort=[(df, -1)])
        result[coll_key] = {
            "first": first.get(df, "") if first else "",
            "last":  last.get(df, "")  if last  else "",
        }

    return jsonify(result)


def _build_former_lookup(coll_key, coll_config, company_id, entity_id, wallet_id, date_to):
    """Return (former_map, sec_dates_map) for the 'previous-date' PU/qty lookup.

    former_map    = {(str(secId), date_str): {"pu", "quantity", "totalContribution"}}
    sec_dates_map = {str(secId): [sorted date strings]}
    All dates <= date_to are included so we can find the immediately preceding date for
    every (security, date) pair in the main query.
    """
    q = {"trashed": {"$ne": True}}
    if company_id:
        q[coll_config["company_field"]] = company_id
    ef = coll_config.get("entity_field")
    if entity_id and ef:
        q[ef] = entity_id
    wf = coll_config.get("wallet_field")
    if wallet_id and wf:
        if isinstance(wallet_id, list):
            q[wf] = {"$in": wallet_id}
        else:
            q[wf] = wallet_id
    df = coll_config["date_field"]
    if date_to:
        q[df] = {"$lte": date_to}

    former_map = {}

    if coll_key == "processedPosition":
        proj = {df: 1, "securities.securityId": 1, "securities.pu": 1,
                "securities.quantity": 1, "securities.totalContribution": 1}
        for doc in db[coll_key].find(q, proj):
            d = str(doc.get(df, "") or "")[:10]
            for sec in (doc.get("securities") or []):
                sid = str(sec.get("securityId", "") or "")
                if sid:
                    former_map[(sid, d)] = {
                        "pu":               sec.get("pu"),
                        "quantity":         sec.get("quantity"),
                        "totalContribution": sec.get("totalContribution"),
                    }
    elif coll_key == "publishedPositionSecurities":
        proj = {df: 1, "securityId": 1, "pu": 1, "quantity": 1, "totalContribution": 1}
        for doc in db[coll_key].find(q, proj):
            d   = str(doc.get(df, "") or "")[:10]
            sid = str(doc.get("securityId", "") or "")
            if sid:
                former_map[(sid, d)] = {
                    "pu":               doc.get("pu"),
                    "quantity":         doc.get("quantity"),
                    "totalContribution": doc.get("totalContribution"),
                }

    # Build sorted date lists per security from the map keys
    sec_dates_map = {}
    for (sid, d) in former_map:
        sec_dates_map.setdefault(sid, []).append(d)
    for sid in sec_dates_map:
        sec_dates_map[sid].sort()

    return former_map, sec_dates_map


def _apply_calcs(rows, calc_keys, former_map, sec_dates_map):
    """Compute requested calc columns in-place; also strips internal _calc_* fields."""
    for row in rows:
        sid    = str(row.pop("_calc_sec_id",        "") or "")
        date   = str(row.pop("_calc_date",          "") or "")[:10]
        cur_pu = row.pop("_calc_pu",                None)
        _      = row.pop("_calc_qty",               None)   # remove even if unused
        cur_tc = row.pop("_calc_total_contrib",     None)

        dates = sec_dates_map.get(sid, [])
        idx   = bisect_left(dates, date) - 1
        if idx >= 0:
            fd        = former_map.get((sid, dates[idx]), {})
            former_pu  = fd.get("pu")
            former_qty = fd.get("quantity")
        else:
            former_pu = former_qty = None

        for calc_key in calc_keys:
            try:
                if calc_key == "calc_rent_pu":
                    row[calc_key] = (cur_pu / former_pu) - 1 if (cur_pu is not None and former_pu) else None
                elif calc_key == "calc_rent_contrib":
                    denom = (former_pu or 0) * (former_qty or 0)
                    row[calc_key] = cur_tc / denom if (cur_tc is not None and denom) else None
                else:
                    row[calc_key] = None
            except (TypeError, ZeroDivisionError):
                row[calc_key] = None


def _resolve_request(body):
    """Parse request body and return (security_ids, lookup_maps, per-collection data)."""
    company_id     = body.get("company_id", "")
    entity_id      = body.get("entity_id",  "")
    wallet_ids_raw = body.get("wallet_ids", [])
    wallet_id      = wallet_ids_raw if wallet_ids_raw else body.get("wallet_id", "")
    security_types = body.get("security_types", [])
    colls_req      = body.get("collections", {})

    security_ids = None
    if security_types:
        security_ids = [
            str(s["_id"]) for s in db.securities.find({"securityType": {"$in": security_types}}, {"_id": 1})
        ]

    lookup_maps = _get_lookup_maps()
    sheets = []

    for coll_key, coll_req in colls_req.items():
        if coll_key not in COLLECTIONS:
            continue
        cfg       = COLLECTIONS[coll_key]
        sel_keys  = coll_req.get("fields", [])
        date_from = coll_req.get("date_from", "")
        date_to   = coll_req.get("date_to",   "")

        if not sel_keys:
            continue

        fm = {f["key"]: f for f in cfg["fields"]}

        # sel_keys may contain '<key>__name' variants for ref fields
        headers  = []
        row_keys = []
        num_keys  = set()
        date_keys = set()
        pct_keys  = set()
        for k in sel_keys:
            is_name = k.endswith("__name")
            base_k  = k[:-6] if is_name else k
            if base_k not in fm:
                continue
            fd = fm[base_k]
            headers.append(fd["label"] + " (Nome)" if is_name else fd["label"])
            row_keys.append(k)
            if not is_name:
                if fd["type"] == "num":  num_keys.add(k)
                if fd["type"] == "date": date_keys.add(k)

        # Calculated columns requested for this collection
        calcs_req   = coll_req.get("calcs", [])
        valid_calcs = [c for c in calcs_req
                       if any(d["key"] == c for d in _CALC_DEFS.get(coll_key, []))]

        query = _build_query(cfg, company_id, entity_id, wallet_id, date_from, date_to, security_ids)
        rows  = _build_rows(coll_key, cfg, sel_keys, query, lookup_maps, security_ids, bool(valid_calcs))

        if valid_calcs:
            former_map, sec_dates_map = _build_former_lookup(
                coll_key, cfg, company_id, entity_id, wallet_id, date_to
            )
            _apply_calcs(rows, valid_calcs, former_map, sec_dates_map)
            for calc_key in valid_calcs:
                calc_def = next(d for d in _CALC_DEFS[coll_key] if d["key"] == calc_key)
                headers.append(calc_def["label"])
                row_keys.append(calc_key)
                num_keys.add(calc_key)
                pct_keys.add(calc_key)

        # Period-based rentability: merge period columns into rows
        period_calcs = [d for d in _CALC_DEFS.get(coll_key, []) if d.get("period_based")]
        if period_calcs:
            nav_rent = body.get("nav_rentability")
            if nav_rent:
                nav_periods   = nav_rent.get("periods", [])
                nav_final     = nav_rent.get("final_date", date_to)
                nav_wids      = nav_rent.get("wallet_ids", []) or (wallet_id if isinstance(wallet_id, list) else ([wallet_id] if wallet_id else []))
                rent_methods  = set(nav_rent.get("methods", ["pu"]))

                if nav_periods and nav_final and nav_wids:
                    period_init_dates = {}
                    for pk in nav_periods:
                        if pk != "since_inception":
                            period_init_dates[pk] = _nav_initial_date(pk, nav_final)

                    coll_sources = set(pc.get("source") for pc in period_calcs)
                    do_nav     = "navPerShare" in coll_sources
                    do_pu      = "pu" in coll_sources and "pu" in rent_methods
                    do_contrib = "contribution" in coll_sources and "contribution" in rent_methods

                    # ── navPerShare (navPackages only) ──
                    if do_nav:
                        final_nav = _nav_latest_by_wallet(nav_wids, nav_final)
                        inception_dates = {}
                        for doc in db.navPackages.aggregate([
                            {"$match": {"walletId": {"$in": nav_wids}, "trashed": {"$ne": True}}},
                            {"$group": {"_id": "$walletId", "minDate": {"$min": "$positionDate"}}},
                        ]):
                            inception_dates[doc["_id"]] = str(doc["minDate"])[:10]

                        init_cache = {}
                        for init_d in set(d for d in period_init_dates.values() if d) | set(d for d in inception_dates.values() if d):
                            init_cache[init_d] = _nav_latest_by_wallet(nav_wids, init_d)

                        rent_by_wallet = {}
                        for wid in nav_wids:
                            fn = final_nav.get(wid)
                            if not fn or not fn.get("navPerShare"):
                                continue
                            final_nps = fn["navPerShare"]
                            rents = {}
                            for pk in nav_periods:
                                init_date = inception_dates.get(wid) if pk == "since_inception" else period_init_dates.get(pk)
                                if init_date and init_date in init_cache:
                                    init_data = init_cache[init_date].get(wid)
                                    if init_data and init_data.get("navPerShare"):
                                        rents[pk] = (final_nps / init_data["navPerShare"]) - 1
                            rent_by_wallet[wid] = rents

                        for row in rows:
                            wid = row.get("walletId") or row.get("_raw_walletId", "")
                            rents = rent_by_wallet.get(wid, {})
                            for pk in nav_periods:
                                row[f"period_rent_{pk}"] = rents.get(pk)

                        for pk in nav_periods:
                            rk = f"period_rent_{pk}"
                            headers.append(_NAV_PERIODS[pk]["label"])
                            row_keys.append(rk)
                            num_keys.add(rk)
                            pct_keys.add(rk)

                    # ── PU point-to-point (position collections) ──
                    if do_pu:
                        # Single query: fetch all PU data once, use bisect for lookups
                        _pu_q = {"walletId": {"$in": nav_wids},
                                 "positionDate": {"$lte": nav_final},
                                 "trashed": {"$ne": True}}
                        if "since_inception" not in nav_periods:
                            _min_init = min((d for d in period_init_dates.values() if d), default=None)
                            if _min_init:
                                _pu_q["positionDate"] = {"$gte": _min_init, "$lte": nav_final}

                        pu_series = {}  # (wid, sid) → sorted [(date_str, pu)]
                        if coll_key in ("processedPosition", "unprocessedSecurityPositions"):
                            for doc in db[coll_key].find(
                                _pu_q, {"walletId": 1, "positionDate": 1,
                                        "securities.securityId": 1, "securities.unprocessedId": 1,
                                        "securities.pu": 1}
                            ).sort("positionDate", 1):
                                d = str(doc.get("positionDate", ""))[:10]
                                wid = str(doc.get("walletId", ""))
                                for sec in doc.get("securities", []):
                                    sid = str(sec.get("securityId", sec.get("unprocessedId", "")))
                                    if sid and sec.get("pu") is not None:
                                        pu_series.setdefault((wid, sid), []).append((d, sec["pu"]))
                        else:
                            for doc in db[coll_key].find(
                                _pu_q, {"walletId": 1, "securityId": 1, "pu": 1, "positionDate": 1}
                            ).sort("positionDate", 1):
                                wid = str(doc.get("walletId", ""))
                                sid = str(doc.get("securityId", ""))
                                if sid and doc.get("pu") is not None:
                                    pu_series.setdefault((wid, sid), []).append(
                                        (str(doc.get("positionDate", ""))[:10], doc["pu"]))

                        # Build parallel date lists for bisect (avoids mixed str/float comparison)
                        pu_dates = {k: [e[0] for e in v] for k, v in pu_series.items()}

                        # Derive inception dates per wallet from fetched data
                        inception_dates = {}
                        for (wid, _sid), entries in pu_series.items():
                            if entries:
                                cur = inception_dates.get(wid)
                                if cur is None or entries[0][0] < cur:
                                    inception_dates[wid] = entries[0][0]

                        def _pu_at(key, target_date):
                            dates = pu_dates.get(key)
                            if not dates:
                                return None
                            idx = bisect_right(dates, target_date) - 1
                            return pu_series[key][idx][1] if idx >= 0 else None

                        for row in rows:
                            wid = row.get("_raw_walletId", "") or row.get("walletId", "")
                            sid = row.get("_raw_securityId", "") or row.get("securityId", "") or row.get("beehusName", "")
                            if not sid:
                                sid = row.get("unprocessedId", "")
                            key = (wid, sid)
                            f_pu = _pu_at(key, nav_final)
                            for pk in nav_periods:
                                rent = None
                                init_date = inception_dates.get(wid) if pk == "since_inception" else period_init_dates.get(pk)
                                if f_pu and init_date:
                                    i_pu = _pu_at(key, init_date)
                                    if i_pu:
                                        rent = (f_pu / i_pu) - 1
                                row[f"period_rent_{pk}"] = rent

                        pu_suffix = " (PU)" if do_contrib else ""
                        for pk in nav_periods:
                            rk = f"period_rent_{pk}"
                            headers.append(_NAV_PERIODS[pk]["label"] + pu_suffix)
                            row_keys.append(rk)
                            num_keys.add(rk)
                            pct_keys.add(rk)

                    # ── Contribution chain-linked (position collections) ──
                    if do_contrib:
                        contrib_inits = dict(period_init_dates)
                        if "since_inception" in nav_periods:
                            contrib_inits["since_inception"] = None
                        contrib_rents = _contrib_period_rent_all(
                            coll_key, nav_wids, contrib_inits, nav_final)

                        for row in rows:
                            wid = row.get("_raw_walletId", "") or row.get("walletId", "")
                            sid = row.get("_raw_securityId", "") or row.get("securityId", "")
                            if not sid:
                                sid = row.get("unprocessedId", "")
                            key = (wid, sid)
                            for pk in nav_periods:
                                row[f"period_contrib_{pk}"] = contrib_rents.get(pk, {}).get(key)

                        contrib_suffix = " (Contrib)" if do_pu else ""
                        for pk in nav_periods:
                            rk = f"period_contrib_{pk}"
                            headers.append(_NAV_PERIODS[pk]["label"] + contrib_suffix)
                            row_keys.append(rk)
                            num_keys.add(rk)
                            pct_keys.add(rk)

            # Strip internal keys
            for row in rows:
                row.pop("_raw_walletId", None)
                row.pop("_raw_securityId", None)

        sheets.append({
            "coll_key":  coll_key,
            "label":     cfg["label"],
            "headers":   headers,
            "row_keys":  row_keys,
            "num_keys":  num_keys,
            "date_keys": date_keys,
            "pct_keys":  pct_keys,
            "rows":      rows,
        })

    return sheets


def _safe(v):
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, datetime):
        return v.isoformat()
    return str(v)


@bp.route("/api/report-builder/saved-reports")
def get_saved_reports():
    return jsonify(_load_saved())


@bp.route("/api/report-builder/saved-reports", methods=["POST"])
def save_report():
    data = request.get_json(force=True)
    reports = _load_saved()
    reports.append({
        "id":         str(uuid.uuid4()),
        "name":       data.get("name", "Sem nome"),
        "created_at": datetime.now().isoformat(),
        "config":     data.get("config", {}),
    })
    _write_saved(reports)
    return jsonify({"ok": True})


@bp.route("/api/report-builder/saved-reports/<rid>", methods=["DELETE"])
def delete_saved_report(rid):
    _write_saved([r for r in _load_saved() if r["id"] != rid])
    return jsonify({"ok": True})


_NAV_PERIODS = {
    "day":   {"label": "Dia",    "months": 0,  "biz_days": 1},
    "month": {"label": "Mês",    "months": 1},
    "3m":    {"label": "3 Meses", "months": 3},
    "6m":    {"label": "6 Meses", "months": 6},
    "12m":   {"label": "12 Meses","months": 12},
    "24m":   {"label": "24 Meses","months": 24},
    "36m":   {"label": "36 Meses","months": 36},
    "year":  {"label": "Ano",    "ytd": True},
    "since_inception": {"label": "Desde Início", "inception": True},
}


def _nav_initial_date(period_key, final_date_str):
    """Return the initial date string for a period relative to final_date."""
    from dateutil.relativedelta import relativedelta
    pdef = _NAV_PERIODS.get(period_key)
    if not pdef:
        return None
    fd = datetime.strptime(final_date_str, "%Y-%m-%d").date()
    if pdef.get("biz_days"):
        from db import get_biz_dates
        dates = get_biz_dates(pdef["biz_days"] + 1, final_date_str)
        return dates[0] if dates else None
    if pdef.get("ytd"):
        return f"{fd.year - 1}-12-31"
    if pdef.get("inception"):
        return None  # handled separately per wallet
    months = pdef.get("months", 0)
    d = fd - relativedelta(months=months)
    return d.strftime("%Y-%m-%d")


def _contrib_period_rent_all(coll_key, wallet_ids, period_init_dates, final_date):
    """Accumulated contribution rent per (walletId, securityId) for multiple periods.

    daily_rent = totalContribution / (former_pu × former_qty)
    period_rent = Π(1 + daily_rent_i) − 1  (chain-linked)

    period_init_dates: {period_key: init_date_str | None}
        None means 'since inception' — accumulate from first available date.
    Returns: {period_key: {(walletId, securityId): rent}}
    """
    all_init = [d for d in period_init_dates.values() if d]
    has_inception = any(v is None for v in period_init_dates.values())

    df = COLLECTIONS[coll_key]["date_field"]
    q = {"walletId": {"$in": wallet_ids}, "trashed": {"$ne": True},
         df: {"$lte": final_date}}
    if all_init and not has_inception:
        q[df] = {"$gte": min(all_init), "$lte": final_date}

    # Build sorted series: (wid, sid) → [(date, tc, pu, qty), …]
    series = {}
    if coll_key == "processedPosition":
        for doc in db[coll_key].find(q, {df: 1, "walletId": 1, "securities": 1}).sort(df, 1):
            d = str(doc.get(df, ""))[:10]
            wid = str(doc.get("walletId", ""))
            for sec in doc.get("securities", []):
                sid = str(sec.get("securityId", ""))
                if sid:
                    series.setdefault((wid, sid), []).append(
                        (d, sec.get("totalContribution"), sec.get("pu"), sec.get("quantity")))
    elif coll_key == "publishedPositionSecurities":
        for doc in db[coll_key].find(
            q, {df: 1, "walletId": 1, "securityId": 1,
                "totalContribution": 1, "pu": 1, "quantity": 1}
        ).sort(df, 1):
            d = str(doc.get(df, ""))[:10]
            wid = str(doc.get("walletId", ""))
            sid = str(doc.get("securityId", ""))
            if sid:
                series.setdefault((wid, sid), []).append(
                    (d, doc.get("totalContribution"), doc.get("pu"), doc.get("quantity")))

    # Pre-compute cumulative products per (wid, sid) — O(S × D)
    cum_map = {}  # key → (dates_list, cum_products)
    for key, entries in series.items():
        dates = [e[0] for e in entries]
        cum = [1.0]  # cum[0] = base (no accumulation)
        for i in range(1, len(entries)):
            tc       = entries[i][1]
            prev_pu  = entries[i - 1][2]
            prev_qty = entries[i - 1][3]
            factor = 1.0
            if tc is not None and prev_pu and prev_qty:
                denom = prev_pu * prev_qty
                if denom:
                    try:
                        factor = 1 + tc / denom
                    except (TypeError, ZeroDivisionError):
                        pass
            cum.append(cum[-1] * factor)
        cum_map[key] = (dates, cum)

    # For each period, use bisect to find start index — O(P × S × log D)
    result = {}
    for pk, init_d in period_init_dates.items():
        period_result = {}
        for key, (dates, cum) in cum_map.items():
            if init_d:
                j = bisect_right(dates, init_d)
            else:
                j = 1  # inception: start from first entry
            if j < 1:
                j = 1
            if j > len(dates):
                period_result[key] = 0.0
                continue
            divisor = cum[j - 1]
            period_result[key] = (cum[-1] / divisor - 1) if divisor else 0.0
        result[pk] = period_result
    return result


def _nav_latest_by_wallet(wallet_ids, max_date):
    """Return {walletId: {navPerShare, date}} for the most recent doc <= max_date per wallet."""
    result = {}
    for doc in db.navPackages.aggregate([
        {"$match": {"walletId": {"$in": wallet_ids}, "positionDate": {"$lte": max_date}, "trashed": {"$ne": True}}},
        {"$sort": {"positionDate": 1}},
        {"$group": {"_id": "$walletId", "nps": {"$last": "$navPerShare"}, "d": {"$last": "$positionDate"}}},
    ]):
        result[doc["_id"]] = {"navPerShare": doc["nps"], "date": str(doc["d"])[:10]}
    return result


@bp.route("/api/report-builder/nav-rentability", methods=["POST"])
def nav_rentability():
    """Compute NAV per-share rentability for wallets across selected periods."""
    data       = request.get_json(force=True)
    wallet_ids = data.get("wallet_ids", [])
    final_date = data.get("final_date", "").strip()
    periods    = data.get("periods", [])

    if not wallet_ids or not final_date or not periods:
        return jsonify({"error": "Parâmetros obrigatórios"}), 400

    # Wallet names (try both string and ObjectId)
    wallet_names = {}
    id_query = list(wallet_ids)
    for wid in wallet_ids:
        try:
            id_query.append(ObjectId(wid))
        except Exception:
            pass
    for w in db.wallets.find({"_id": {"$in": id_query}}, {"name": 1}):
        wallet_names[str(w["_id"])] = w.get("name", "")

    # Final navPerShare per wallet (single aggregation)
    final_nav = _nav_latest_by_wallet(wallet_ids, final_date)

    # Inception dates (single aggregation)
    inception_dates = {}
    for doc in db.navPackages.aggregate([
        {"$match": {"walletId": {"$in": wallet_ids}, "trashed": {"$ne": True}}},
        {"$group": {"_id": "$walletId", "minDate": {"$min": "$positionDate"}}},
    ]):
        inception_dates[doc["_id"]] = str(doc["minDate"])[:10]

    # Pre-compute initial dates per period (shared across all wallets except inception)
    period_init_dates = {}
    for pk in periods:
        if pk != "since_inception":
            period_init_dates[pk] = _nav_initial_date(pk, final_date)

    # Batch: fetch initial navPerShare for all periods in one pass per unique date
    # Group periods by their init_date to minimize queries
    date_to_periods = {}  # {init_date: [period_keys]}
    for pk, d in period_init_dates.items():
        if d:
            date_to_periods.setdefault(d, []).append(pk)
    # For inception, each wallet has a different date — collect all unique dates
    inception_unique = set(inception_dates.values())
    for d in inception_unique:
        if d:
            date_to_periods.setdefault(d, [])  # ensure we query it

    # One aggregation per unique init_date → {walletId: {navPerShare, date}}
    init_cache = {}  # {init_date: {walletId: {navPerShare, date}}}
    for init_d in date_to_periods:
        init_cache[init_d] = _nav_latest_by_wallet(wallet_ids, init_d)

    # Build results
    rows = []
    for wid in wallet_ids:
        fn = final_nav.get(wid)
        if not fn:
            continue
        final_nps = fn["navPerShare"]
        row = {"walletId": wid, "walletName": wallet_names.get(wid, wid),
               "finalDate": fn["date"], "finalNavPerShare": final_nps}

        for pk in periods:
            if pk == "since_inception":
                init_date = inception_dates.get(wid)
            else:
                init_date = period_init_dates.get(pk)

            rent = None
            if init_date and init_date in init_cache:
                init_data = init_cache[init_date].get(wid)
                if init_data and init_data.get("navPerShare"):
                    init_nps = init_data["navPerShare"]
                    rent = (final_nps / init_nps) - 1 if init_nps else None
                    row[f"{pk}_initialDate"] = init_data["date"]
                    row[f"{pk}_initialNavPerShare"] = init_nps

            row[pk] = rent

        rows.append(row)

    return jsonify({"rows": rows, "periods": periods, "periodLabels": {k: _NAV_PERIODS[k]["label"] for k in periods}})


@bp.route("/api/report-builder/preview", methods=["POST"])
def preview():
    sheets = _resolve_request(request.get_json(force=True))
    result = []
    for s in sheets:
        rows_preview = [
            [_safe(row.get(k)) for k in s["row_keys"]]
            for row in s["rows"][:200]
        ]
        # Mark which column indices are percentages / numbers
        pct_indices = [i for i, k in enumerate(s["row_keys"]) if k in s.get("pct_keys", set())]
        num_indices = [i for i, k in enumerate(s["row_keys"]) if k in s.get("num_keys", set())]
        result.append({
            "label":      s["label"],
            "headers":    s["headers"],
            "rows":       rows_preview,
            "total":      len(s["rows"]),
            "pctIndices": pct_indices,
            "numIndices": num_indices,
        })
    return jsonify(result)


@bp.route("/api/report-builder/send-email", methods=["POST"])
def send_email_report():
    data        = request.get_json(force=True)
    to_list     = [e.strip() for e in data.get("to", []) if e.strip()]
    subject     = data.get("subject", "Relatório")
    email_from  = os.environ.get("EMAIL_FROM", "tecnologia@beehus.com.br")
    email_pass  = os.environ.get("EMAIL_PASSWORD")

    if not to_list:
        return jsonify({"ok": False, "error": "Nenhum destinatário informado."}), 400
    if not email_pass:
        return jsonify({"ok": False, "error": "Variável EMAIL_PASSWORD não configurada no servidor."}), 500

    sheets = _resolve_request(data.get("report_config", {}))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for s in sheets:
        ws = wb.create_sheet(title=s["label"][:31])
        _write_sheet(ws, s["headers"], s["row_keys"], s["rows"], s["num_keys"], s.get("date_keys"), s.get("pct_keys"))
    if not wb.worksheets:
        wb.create_sheet("Sem dados")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    msg = MIMEMultipart()
    msg["From"]    = email_from
    msg["To"]      = ", ".join(to_list)
    msg["Subject"] = subject
    msg.attach(MIMEText("Segue em anexo o relatório solicitado.", "plain", "utf-8"))

    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(buf.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename="relatorio.xlsx")
    msg.attach(part)

    try:
        with smtplib.SMTP("smtp.office365.com", 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(email_from, email_pass)
            smtp.sendmail(email_from, to_list, msg.as_string())
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@bp.route("/api/report-builder/generate", methods=["POST"])
def generate():
    sheets = _resolve_request(request.get_json(force=True))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for s in sheets:
        ws = wb.create_sheet(title=s["label"][:31])
        _write_sheet(ws, s["headers"], s["row_keys"], s["rows"], s["num_keys"], s.get("date_keys"), s.get("pct_keys"))

    if not wb.worksheets:
        wb.create_sheet("Sem dados")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        download_name="relatorio.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ══════════════════════════════════════════════════════════════════════════════
# DERIVATIVES
# ══════════════════════════════════════════════════════════════════════════════

def _load_deriv_mappings():
    """Load persistent txn→position mappings. {txnId: {unprocessedId, buyPU, include}}"""
    if not os.path.exists(DERIV_MAPPINGS_FILE):
        return {}
    with open(DERIV_MAPPINGS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_deriv_mappings(mappings):
    with open(DERIV_MAPPINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(mappings, f, indent=2, ensure_ascii=False)


def _resolve_names(ids, collection, field="beehusName"):
    """Return {str(id): name} for a set of string IDs."""
    oids = []
    for s in ids:
        try:
            oids.append(ObjectId(s))
        except Exception:
            pass
    if not oids:
        return {}
    return {str(d["_id"]): d.get(field, "") for d in db[collection].find({"_id": {"$in": oids}}, {field: 1})}


@bp.route("/api/report-builder/deriv/transactions", methods=["POST"])
def deriv_transactions():
    """Return all transactions up to selected date, enriched with stored mappings."""
    body = request.get_json(force=True)
    wallet_ids = body.get("wallet_ids", [])
    date_str   = body.get("date", "")
    if not wallet_ids or not date_str:
        return jsonify([])

    q = {
        "walletId":        {"$in": wallet_ids},
        "liquidationDate": date_str,
        "trashed":         {"$ne": True},
    }
    sec_ids = set()
    txns = []
    for doc in db.transactions.find(q, {
        "walletId": 1, "securityId": 1, "quantity": 1,
        "price": 1, "balance": 1, "description": 1,
        "operationDate": 1, "liquidationDate": 1,
    }).sort("liquidationDate", -1):
        sid = str(doc.get("securityId", "") or "")
        sec_ids.add(sid)
        txns.append({
            "id":              str(doc["_id"]),
            "walletId":        str(doc.get("walletId", "")),
            "securityId":      sid,
            "quantity":        doc.get("quantity"),
            "price":           doc.get("price"),
            "balance":         doc.get("balance"),
            "description":     doc.get("description", ""),
            "operationDate":   str(doc.get("operationDate", ""))[:10],
            "liquidationDate": str(doc.get("liquidationDate", ""))[:10],
        })

    sec_map    = _resolve_names(sec_ids, "securities", "beehusName")
    wallet_map = _resolve_names(wallet_ids, "wallets", "name")

    # Enrich with stored mappings
    stored = _load_deriv_mappings()

    for t in txns:
        t["securityName"] = sec_map.get(t["securityId"], "")
        t["walletName"]   = wallet_map.get(t["walletId"], "")
        m = stored.get(t["id"])
        if m:
            t["storedMapping"] = {
                "unprocessedId": m.get("unprocessedId", ""),
                "buyPU":         m.get("buyPU", 0),
                "include":       m.get("include", True),
            }

    return jsonify(txns)


@bp.route("/api/report-builder/deriv/positions", methods=["POST"])
def deriv_positions():
    """Return unprocessedSecurityPositions.securities[] with securityType resolved via mappings."""
    body = request.get_json(force=True)
    wallet_ids = body.get("wallet_ids", [])
    date_str   = body.get("date", "")
    company_id = body.get("company_id", "")
    if not wallet_ids or not date_str:
        return jsonify([])

    # Build unprocessedId → securityId mapping
    mapping = _build_security_mapping(company_id) if company_id else {}

    q = {
        "walletId":     {"$in": wallet_ids},
        "positionDate": date_str,
        "trashed":      {"$ne": True},
    }
    positions = []
    all_sec_ids = set()
    for doc in db.unprocessedSecurityPositions.find(q, {
        "walletId": 1, "securities": 1,
    }):
        wid = str(doc.get("walletId", ""))
        for sec in doc.get("securities", []):
            uid = str(sec.get("unprocessedId", ""))
            sid = str(sec.get("securityId", "")) or uid
            # Resolve via mapping: unprocessedId → securityId
            mapped_sid = mapping.get(uid, "")
            all_sec_ids.add(sid)
            if mapped_sid:
                all_sec_ids.add(mapped_sid)
            positions.append({
                "walletId":      wid,
                "securityId":    sid,
                "unprocessedId": uid,
                "mappedSecId":   mapped_sid,
                "quantity":      sec.get("quantity"),
                "pu":            sec.get("pu"),
                "amount":        sec.get("amount"),
            })

    # Resolve names + securityType from securities collection
    sec_info = {}
    if all_sec_ids:
        oids = []
        for s in all_sec_ids:
            try:
                oids.append(ObjectId(s))
            except Exception:
                pass
        for s in db.securities.find({"_id": {"$in": oids}}, {"beehusName": 1, "securityType": 1}):
            sec_info[str(s["_id"])] = {
                "name": s.get("beehusName", ""),
                "type": s.get("securityType", ""),
            }

    # Former quantities: most recent unprocessedSecurityPositions before date_str
    former_qty = {}  # (walletId, unprocessedId) → quantity
    for doc in db.unprocessedSecurityPositions.find(
        {"walletId": {"$in": wallet_ids}, "positionDate": {"$lt": date_str}, "trashed": {"$ne": True}},
        {"walletId": 1, "positionDate": 1, "securities.unprocessedId": 1, "securities.quantity": 1},
    ).sort("positionDate", -1):
        wid = str(doc.get("walletId", ""))
        for sec in doc.get("securities", []):
            uid = str(sec.get("unprocessedId", ""))
            key = (wid, uid)
            if key not in former_qty and sec.get("quantity") is not None:
                former_qty[key] = sec["quantity"]

    # Stored buyPU: reverse-lookup from deriv_mappings (txnId→{unprocessedId,buyPU})
    # Build (walletId, unprocessedId) → buyPU using transaction walletId
    stored_buy_pu = {}  # (walletId, unprocessedId) → buyPU
    stored_mappings = _load_deriv_mappings()
    if stored_mappings:
        # Fetch walletId for each mapped transaction
        txn_oids = []
        for tid in stored_mappings:
            try:
                txn_oids.append(ObjectId(tid))
            except Exception:
                pass
        txn_wallets = {}
        if txn_oids:
            for doc in db.transactions.find({"_id": {"$in": txn_oids}}, {"walletId": 1}):
                txn_wallets[str(doc["_id"])] = str(doc.get("walletId", ""))
        for tid, m in stored_mappings.items():
            wid = txn_wallets.get(tid, "")
            uid = m.get("unprocessedId", "")
            bpu = m.get("buyPU", 0)
            incl = m.get("include", True)
            if wid and uid and bpu:
                stored_buy_pu[(wid, uid)] = {"buyPU": bpu, "include": incl}

    for p in positions:
        info = sec_info.get(p["securityId"]) or sec_info.get(p["mappedSecId"]) or {}
        p["securityName"] = info.get("name") or p["unprocessedId"]
        p["securityType"] = info.get("type", "")
        p["mapped"] = bool(p["mappedSecId"] and sec_info.get(p["mappedSecId"]))
        p["formerQuantity"] = former_qty.get((p["walletId"], p["unprocessedId"]))
        sbpu = stored_buy_pu.get((p["walletId"], p["unprocessedId"]))
        if sbpu:
            p["storedBuyPU"] = sbpu["buyPU"]
            p["storedInclude"] = sbpu["include"]

    return jsonify(positions)


@bp.route("/api/report-builder/deriv/mappings", methods=["POST"])
def save_deriv_mappings():
    """Save persistent transaction→position mappings.
    Body: {mappings: [{txnId, unprocessedId, buyPU, include}, ...]}
    """
    body = request.get_json(force=True)
    new_items = body.get("mappings", [])
    stored = _load_deriv_mappings()
    for item in new_items:
        tid = item.get("txnId")
        if not tid:
            continue
        stored[tid] = {
            "unprocessedId": item.get("unprocessedId", ""),
            "buyPU":         item.get("buyPU", 0),
            "include":       item.get("include", True),
        }
    _save_deriv_mappings(stored)
    return jsonify({"ok": True, "count": len(stored)})


@bp.route("/api/report-builder/deriv/mappings")
def get_deriv_mappings():
    """Return all stored mappings enriched with transaction info."""
    stored = _load_deriv_mappings()
    if not stored:
        return jsonify([])

    # Fetch transaction details for display
    oids = []
    for tid in stored:
        try:
            oids.append(ObjectId(tid))
        except Exception:
            pass
    txn_info = {}
    if oids:
        for doc in db.transactions.find({"_id": {"$in": oids}}, {
            "walletId": 1, "description": 1, "quantity": 1,
            "balance": 1, "liquidationDate": 1,
        }):
            txn_info[str(doc["_id"])] = {
                "walletId":        str(doc.get("walletId", "")),
                "description":     doc.get("description", ""),
                "quantity":        doc.get("quantity"),
                "balance":         doc.get("balance"),
                "liquidationDate": str(doc.get("liquidationDate", ""))[:10],
            }

    # Wallet names
    wids = list(set(t.get("walletId", "") for t in txn_info.values()))
    wallet_map = _resolve_names(wids, "wallets", "name")

    result = []
    for tid, m in stored.items():
        ti = txn_info.get(tid, {})
        result.append({
            "txnId":           tid,
            "unprocessedId":   m.get("unprocessedId", ""),
            "buyPU":           m.get("buyPU", 0),
            "include":         m.get("include", True),
            "description":     ti.get("description", ""),
            "quantity":        ti.get("quantity"),
            "balance":         ti.get("balance"),
            "liquidationDate": ti.get("liquidationDate", ""),
            "walletName":      wallet_map.get(ti.get("walletId", ""), ""),
        })
    result.sort(key=lambda r: r.get("liquidationDate", ""), reverse=True)
    return jsonify(result)


@bp.route("/api/report-builder/deriv/mappings/<tid>", methods=["DELETE"])
def delete_deriv_mapping(tid):
    stored = _load_deriv_mappings()
    stored.pop(tid, None)
    _save_deriv_mappings(stored)
    return jsonify({"ok": True})
